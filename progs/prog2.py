from datetime import datetime
from decimal import Decimal
from konfs.konfs import SH

from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill, Border, Side
import openpyxl, time, os


def prog2(fullpathtoxlsxfile):
    start_time = time.time()
    wb = openpyxl.load_workbook(fullpathtoxlsxfile)
    # wb = Workbook()
    ws = wb.worksheets[1]

    DTVtable=[[ws.cell(row=i, column=j).value for j in range(1, ws.max_column+1)] for i in range(2, ws.max_row+1)]
    wb.close()
    # корректируем столбец с нумерацией
    for i in range(0, (len(DTVtable))):
        DTVtable[i][0] = i+1

    Jtable=[[DTVtable[i][j] for j in (4,1,2,5)] for i in range(0, (len(DTVtable)))]
    # print(Jtable)
    for i in range(0, (len(Jtable))):
        Jtable[i].insert(1, (Jtable[i][0]).date())
    # print(Jtable)
    Jtable=[[Jtable[i][j] for j in (1,2,3,4)] for i in range(0, (len(Jtable)))]
    # Создаем список Jtable_out (суммируется в течение дня)
    key = lambda item: (item[0], item[1], item[2])
    keys = set(map(key, Jtable))
    Jtable_out = [[k[0], k[1], k[2], sum(Decimal(n[3]) for n in Jtable if k == key(n))] for k in keys]
    Jtable=Jtable_out
    # сортируем по возрастанию даты
    Jtable.sort(key=lambda x: x[0])
    # добавляем столбец с нумерацией
    m=str(Jtable[-1][0].month)
    for i in range(0, (len(Jtable))):
        Jtable[i].insert(0, i+1)
        Jtable[i].insert(-1, f'{m}-{str(Jtable[i][0])}')
        Jtable[i].insert(-1, SH)
    for i in range(0, (len(Jtable))):
        Jtable[i][-1]=(Jtable[i][-1]/1000)
    # формируем список по форме журнала:
    Jtable=[[Jtable[i][j] for j in (0,1,2,3,6,6,4,1,5,6,6)] for i in range(0, (len(Jtable))) if Jtable[i][6] != 0]
    # добавляем столбец с нумерацией
    for i in range(0, (len(Jtable))):
        Jtable[i][0] = i+1
    for i in range(0, (len(Jtable))):
        Jtable[i][6]=f'{m}-{str(Jtable[i][0])}'
    sum_Jtable=sum([Jtable[i][4] for i in range(0, (len(Jtable)))])
    # print(sum_Jtable)
    Jtable_itog=[["", "", "", "Итого за месяц:", sum_Jtable,sum_Jtable,"","","",sum_Jtable,sum_Jtable,"","","","","","", "0,000", "0,000"]]

    wb = openpyxl.load_workbook(r"..\konfs\Шаблоны\Шаблон.xlsx")
    ws0 = wb.worksheets[0]
    ws = wb.worksheets[1]

    for n, row in enumerate(Jtable, 10):
        ws0.append(row)
        for i in (1,5,6,7,9,10,11):
            ws0.cell(row=n, column=i).style = "style9c"
        for i in (3,4):
            ws0.cell(row=n, column=i).style = "style9"
        for i in (2, 8):
            ws0.cell(row=n, column=i).style = "datestyle_short"
        for i in range(1, 22):
            Thin=Side(border_style='thin', color="FF000000")
            ws0.cell(row=n, column=i).border = Border(top=Thin, bottom=Thin, left=Thin, right=Thin)
    lr=ws0.max_row
    for row in Jtable_itog:
        ws0.append(row)
        for i in (5, 6, 10, 11, 18,19):
            ws0.cell(row=lr+1, column=i).style = "style9cb"
        for i in (3,4):
            ws0.cell(row=lr+1, column=i).style = "style9rb"
        for i in range(1, 22):
            Thin = Side(border_style='thin', color="FF000000")
            ws0.cell(row=lr+1, column=i).border = Border(top=Thin, bottom=Thin, left=Thin, right=Thin)

    for n, row in enumerate(DTVtable, 2):
        ws.append(row)
        for i in (1, 4, 14, 15):
            ws.cell(row=n, column=i).style = "style11"
        for i in (2, 3):
            ws.cell(row=n, column=i).style = "style11l"
        for i in (6, 13):
            ws.cell(row=n, column=i).style = "style16"
        for i in (7,8,9):
            ws.cell(row=n, column=i).style = "style9"
        for i in (5,10,11,12):
            ws.cell(row=n, column=i).style = "datestyle"
        for i in range(1, 16):
            if divmod(n, 2)[1] == 0:
                ws.cell(row=n, column=i).fill = PatternFill('solid', fgColor="D9D9D9")
        for i in range(1, 16):
            badcell=ws.cell(row=n, column=i).value
            if badcell == 'Станция не определена' or badcell == datetime.strptime("00:00 01.01.01", "%H:%M %d.%m.%y") or \
            badcell == 'Не определено наименование участка МН' or badcell == 'Кол-во твердого не определено' :
                for j in range(1, 16):
                    ws.cell(row=n, column=j).fill = PatternFill('solid', fgColor="D97147")
        # print(n)
        ws.delete_cols(16)
    tab = Table(displayName="VTD", ref=f"A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}")
    # print(ws['E2'].value)
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 55
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 15
    ws.column_dimensions['M'].width = 10
    ws.column_dimensions['N'].width = 17
    ws.column_dimensions['O'].width = 42

    style = TableStyleInfo(name="TableStyleLight9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style
    # сохраняем чтобы перенести таблицу в новый файл из файла шаблона, чтобы не было ошибки задвоения имени таблицы
    wb.save(r"..\2_Журнал_и_акты\Журнал_НШЛ.xlsx")
    ws.add_table(tab)
    wb.save(r"..\2_Журнал_и_акты\Журнал_НШЛ.xlsx")

    print("Затрачено времени: %s секунд " % (time.time() - start_time))
    b=os.path.abspath(r"..\2_Журнал_и_акты\Журнал_НШЛ.xlsx")
    return b

if __name__ == '__main__':
    print(prog2(r"..\2_Журнал_и_акты\Журнал_НШЛ.xlsx"))